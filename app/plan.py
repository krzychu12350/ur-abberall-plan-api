import json

import openpyxl
import re

# Load the Excel workbook
workbook = openpyxl.load_workbook('plan.xlsx')
sheet = workbook['Arkusz1']  # Replace 'Arkusz1' with your actual sheet name
data_list = []


def remove_words_from_string(input_string, words_to_remove):
    # Split the input string into words
    words = input_string.split()
    # Create a new list that excludes words in the words_to_remove array
    filtered_words = [word for word in words if word not in words_to_remove]
    # Join the filtered words back into a string
    return ' '.join(filtered_words)


def find_nearest_godzina(col, start_row):
    """Find the nearest cell with value 'GODZINA' above the specified row in the given column."""
    for row in range(start_row - 1, 0, -1):  # Go upwards from start_row
        cell_value = sheet.cell(row=row, column=col).value
        if isinstance(cell_value, str) and cell_value.strip() == 'GODZINA':
            return row
    return None


# Function to extract and format the date
# Regular expression to match only the date (ignore the day of the week)
pattern = r'\b\w+\s(\d{1,2})\.(\d{1,2})\.(\d{4})\b'


# Function to extract and format the date
def extract_date(text):
    match = re.search(pattern, text)
    if match:
        day, month, year = match.groups()
        # Format the date with leading zeros for day and month
        formatted_date = f"{int(day):02d}.{int(month):02d}.{year}"
        return formatted_date
    return "No date found"


def extract_data():
    # Iterate through merged cell ranges
    for merged_range in sheet.merged_cells.ranges:

        # Skip rows 1 to 3 and rows 947 till the end
        if merged_range.bounds[1] <= 3 or merged_range.bounds[1] >= 947:
            continue

        # Get the number of cells in the merged range
        min_row, min_col, max_row, max_col = merged_range.bounds

        num_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        # print(merged_range.bounds)

        # Access the second and fourth elements
        second_element = merged_range.bounds[1]  # Index 1 for the second element
        fourth_element = merged_range.bounds[3]  # Index 3 for the fourth element
        number_of_rows = fourth_element - second_element

        if number_of_rows < 1:
            continue

        # Access the value of the top-left cell of the merged range
        min_value_cell_data = sheet.cell(row=merged_range.bounds[1], column=merged_range.bounds[0]).value

        # Check if the cell data is a string
        if isinstance(min_value_cell_data, str):
            # Remove excessive whitespace using regex
            cleaned_text = re.sub(r'\s+', ' ', min_value_cell_data).strip()
        else:
            cleaned_text = str(min_value_cell_data)

            # Assign time ranges from column A for each row in the merged range
        time_ranges = []
        for row in range(second_element, fourth_element + 1):
            time_range = sheet.cell(row=row, column=1).value  # Assuming column A is the first column
            time_ranges.append(time_range)

        if time_ranges[0].split(' - ')[0] == 'GODZINA':
            continue
        else:
            # print(time_ranges)
            # start_time = time_ranges[0].split(' - ')[0]
            # end_time = time_ranges[-1].split(' - ')[1]

            # Splitting the first and last elements
            start_time = re.split(r' - |- ', time_ranges[0])[0]
            end_time = re.split(r' - |- ', time_ranges[-1])[-1]
            # print(time_ranges[0].split(' - ')[0])
            # print(time_ranges[-1].split(' - ')[1])

        nearest_godzina_row = find_nearest_godzina(1, merged_range.bounds[1])

        # column
        # print(merged_range.bounds[0])

        # row
        # print(nearest_godzina_row)
        # if merged_range.bounds[0] > 2 and merged_range.bounds[0] <= 6:
        #     date_column = 3
        # elif merged_range.bounds[0] >= 7 and merged_range.bounds[0] <= 10:
        #     date_column = 7
        # elif merged_range.bounds[0] >= 11 and merged_range.bounds[0] <= 14:
        #     date_column = 11
        # elif merged_range.bounds[0] >= 15 and merged_range.bounds[0] <= 18:
        #     date_column = 15
        # elif merged_range.bounds[0] >= 19 and merged_range.bounds[0] <= 22:
        #     date_column = 19

        date_column = 0
        if 2 < merged_range.bounds[0] <= 22:
            date_column = 3 + 4 * ((merged_range.bounds[0] - 3) // 4)

        date = sheet.cell(row=nearest_godzina_row, column=date_column).value

        words_to_remove = ["PONIEDZIAŁEK", "WTOREK", "ŚRODA", "CZWARTEK", "PIĄTEK"]

        fromated_date = remove_words_from_string(date, words_to_remove)

        # print(extract_date(date))  # Output: 06.12.2024

        # print(f'Merged Range: {merged_range}, Number of Rows: { number_of_rows}, Min Value Cell Data: {cleaned_text}, Start time: {start_time}, End_time: {end_time}')

        data = {
            'text': cleaned_text,
            'date': fromated_date,
            'start_time': start_time,
            'end_time': end_time,
            # 'nearest_godzina_row': nearest_godzina_row,
            # 'nearest_godzina_column': merged_range.bounds[0],
            # 'lesson_block_cells_range': merged_range.bounds,

        }
        # Append each data object to the list
        data_list.append(data)
        # print('----------------------------')

    # Convert the list of dictionaries to a JSON array of objects
    # json_data = json.dumps(data_list, indent=4, ensure_ascii=False)

    # Print the resulting JSON array
    # print(json_data)
    return data_list

