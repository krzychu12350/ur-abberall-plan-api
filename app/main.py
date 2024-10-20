import mysql.connector
from fastapi import FastAPI, __version__, HTTPException, UploadFile
from fastapi.responses import JSONResponse
from fastapi import FastAPI, UploadFile, File, HTTPException
from collections import defaultdict
from datetime import datetime, timedelta
from typing import List, Dict, Any
import cloudinary
import cloudinary.uploader
import cloudinary.api
import os
from dotenv import load_dotenv
from fastapi.middleware.cors import CORSMiddleware
import json
from datetime import datetime
import openpyxl
import re
import openpyxl
import re
from datetime import datetime
import requests
from io import BytesIO
import requests
import pdfplumber
import pandas as pd
from io import BytesIO

# Lo
# ad environment variables from .env file
load_dotenv()

# Initialize Cloudinary
cloudinary.config(
    cloud_name=os.getenv('CLOUDINARY_CLOUD_NAME'),
    api_key=os.getenv('CLOUDINARY_API_KEY'),
    api_secret=os.getenv('CLOUDINARY_API_SECRET')
)

import openpyxl
import re
from datetime import datetime

# Precompile the regular expression for date extraction
pattern = re.compile(r'\b\w+\s(\d{1,2})\.(\d{1,2})\.(\d{4})\b')

# URL of the Excel file
EXCEL_FILE_URL = "http://ur.edu.pl/files/user_directory/307/RAT%20MED%201%20ROK%202024%20stacjonarne.xlsx"

# MySQL connection configuration
DB_CONFIG = {
    'host': os.getenv('MYSQL_HOST', 'localhost'),
    'user': os.getenv('MYSQL_USER', 'root'),
    'password': os.getenv('MYSQL_PASSWORD', ''),
    'database': os.getenv('MYSQL_DB', 'schedules_db')
}


def get_db_connection():
    return mysql.connector.connect(**DB_CONFIG)


def split_time_range(time_range):
    if ' - ' in time_range:
        return time_range.split(' - ')
    elif '- ' in time_range:
        return time_range.split('- ')
    else:
        return None, None  # In case the format is unexpected


# Optimize the word removal by using a set for faster lookup
def remove_words_from_string(input_string, words_to_remove):
    words_to_remove_set = set(words_to_remove)
    return ' '.join(word for word in input_string.split() if word not in words_to_remove_set)


def find_nearest_godzina(sheet, col, start_row):
    """Find the nearest cell with value 'GODZINA' above the specified row in the given column."""
    for row in range(start_row - 1, 0, -1):  # Go upwards from start_row
        cell_value = sheet.cell(row=row, column=col).value
        if isinstance(cell_value, str) and cell_value.strip() == 'GODZINA':
            return row
    return None


def extract_date(text):
    """Extract and format the date from text using a precompiled regex."""
    match = pattern.search(text)
    if match:
        day, month, year = match.groups()
        return f"{int(day):02d}.{int(month):02d}.{year}"
    return "No date found"


def parse_time(time_str: str) -> datetime:
    """Parses a time string into a datetime object. Handles time ranges by returning the start time."""
    if '-' in time_str:
        time_str = time_str.split('-')[0].strip()  # Take the first part as the start time
    return datetime.strptime(time_str.strip(), '%H:%M')


def extract_data():
    # Download the Excel file from the provided URL
    response = requests.get(EXCEL_FILE_URL)

    if response.status_code == 200:
        # Load the Excel file into memory
        file_stream = BytesIO(response.content)
        workbook = openpyxl.load_workbook(file_stream)
    else:
        raise Exception(f"Failed to download the file. Status code: {response.status_code}")

    sheet = workbook['Arkusz1']  # Replace 'Arkusz1' with your actual sheet name

    data_list = []
    words_to_remove = ["PONIEDZIAŁEK", "WTOREK", "ŚRODA", "CZWARTEK", "PIĄTEK"]

    # Iterate through merged cell ranges
    for merged_range in sheet.merged_cells.ranges:

        # Skip rows 1 to 3 and rows 947 till the end
        if merged_range.bounds[1] <= 3 or merged_range.bounds[1] >= 947:
            continue

        # Get the number of cells in the merged range
        min_row, min_col, max_row, max_col = merged_range.bounds

        num_cells = (max_row - min_row + 1) * (max_col - min_col + 1)
        # print(merged_range.bounds)

        # Access the second and fourth elements
        second_element = merged_range.bounds[1]  # Index 1 for the second element
        fourth_element = merged_range.bounds[3]  # Index 3 for the fourth element
        number_of_rows = fourth_element - second_element

        if number_of_rows < 1:
            continue

        # Access the value of the top-left cell of the merged range
        min_value_cell_data = sheet.cell(row=merged_range.bounds[1], column=merged_range.bounds[0]).value

        # Check if the cell data is a string
        if isinstance(min_value_cell_data, str):
            # Remove excessive whitespace using regex
            cleaned_text = re.sub(r'\s+', ' ', min_value_cell_data).strip()
        else:
            cleaned_text = str(min_value_cell_data)

            # Assign time ranges from column A for each row in the merged range
        time_ranges = []
        for row in range(second_element, fourth_element + 1):
            time_range = sheet.cell(row=row, column=1).value  # Assuming column A is the first column
            time_ranges.append(time_range)

        if time_ranges[0].split(' - ')[0] == 'GODZINA':
            continue
        else:
            # print(time_ranges)
            # start_time = time_ranges[0].split(' - ')[0]
            # end_time = time_ranges[-1].split(' - ')[1]
            # Handling start_time and end_time based on input ranges
            start_range = time_ranges[0]
            end_range = time_ranges[-1]

            # Splitting start_time and end_time based on the delimiter found
            start_time, _ = split_time_range(start_range)  # Only need the start part
            _, end_time = split_time_range(end_range)  # Only need the end part
            # Splitting the first and last elements
            # start_time = re.split(r' - |- ', time_ranges[0])[0]
            # end_time = re.split(r' - |- ', time_ranges[-1])[-1]
            # print(time_ranges[0].split(' - ')[0])
            # print(time_ranges[-1].split(' - ')[1])

        nearest_godzina_row = find_nearest_godzina(sheet, 1, merged_range.bounds[1])

        # column
        # print(merged_range.bounds[0])

        # row
        # print(nearest_godzina_row)
        # if merged_range.bounds[0] > 2 and merged_range.bounds[0] <= 6:
        #     date_column = 3
        # elif merged_range.bounds[0] >= 7 and merged_range.bounds[0] <= 10:
        #     date_column = 7
        # elif merged_range.bounds[0] >= 11 and merged_range.bounds[0] <= 14:
        #     date_column = 11
        # elif merged_range.bounds[0] >= 15 and merged_range.bounds[0] <= 18:
        #     date_column = 15
        # elif merged_range.bounds[0] >= 19 and merged_range.bounds[0] <= 22:
        #     date_column = 19

        date_column = 0
        if 2 < merged_range.bounds[0] <= 22:
            date_column = 3 + 4 * ((merged_range.bounds[0] - 3) // 4)

        date = sheet.cell(row=nearest_godzina_row, column=date_column).value

        words_to_remove = ["PONIEDZIAŁEK", "WTOREK", "ŚRODA", "CZWARTEK", "PIĄTEK"]

        fromated_date = remove_words_from_string(date, words_to_remove)

        # print(extract_date(date))  # Output: 06.12.2024

        # print(f'Merged Range: {merged_range}, Number of Rows: { number_of_rows}, Min Value Cell Data: {cleaned_text}, Start time: {start_time}, End_time: {end_time}')

        data = {
            'text': cleaned_text,
            'date': fromated_date,
            'start_time': start_time,
            'end_time': end_time,
            # 'nearest_godzina_row': nearest_godzina_row,
            # 'nearest_godzina_column': merged_range.bounds[0],
            # 'lesson_block_cells_range': merged_range.bounds,

        }
        # Append each data object to the list
        data_list.append(data)
        # print('----------------------------')

    # Convert the list of dictionaries to a JSON array of objects
    # json_data = json.dumps(data_list, indent=4, ensure_ascii=False)

    # Print the resulting JSON array
    # print(json_data)

    # Sort the data list by start_time in ascending order
    data_list.sort(key=lambda x: parse_time(x['start_time']))
    return data_list


def save_data_to_db(data_list):
    connection = get_db_connection()
    cursor = connection.cursor()

    # Truncate the table before inserting new data
    cursor.execute("TRUNCATE TABLE schedules")

    for data in data_list:
        # Convert the date from 'DD.MM.YYYY' to 'YYYY-MM-DD'
        date_object = datetime.strptime(data['date'], '%d.%m.%Y')  # Parse the date string
        formatted_date = date_object.strftime('%Y-%m-%d')  # Format the date
        # print(formatted_date)
        query = """
        INSERT INTO schedules (text, date, start_time, end_time)
        VALUES (%s, %s, %s, %s)
        """
        # print(f"Inserting data for date: {formatted_date}")  # Log the formatted date
        cursor.execute(query, (data['text'], formatted_date, data['start_time'], data['end_time']))

    connection.commit()
    cursor.close()
    connection.close()


origins = ["*"]
app = FastAPI()
app.add_middleware(
    CORSMiddleware,
    allow_origins=origins,
    allow_credentials=True,
    allow_methods=["*"],
    allow_headers=["*"],
)


# app.include_router(system.router, prefix="/system")


@app.get("/statuss")
def status():
    return {"ok": True, "version": __version__}


@app.get("/")
def status():
    return {"ok": True, 'page': 'home'}


@app.get("/api/schedules", response_model=Dict[str, List[Dict[str, Any]]])
async def get_schedules():
    data = extract_data()

    # Group schedules by date
    grouped_schedules = defaultdict(list)

    for schedule in data:
        grouped_schedules[schedule['date']].append(schedule)

    # Sort the grouped schedules by date
    sorted_grouped_schedules = dict(
        sorted(grouped_schedules.items(), key=lambda x: datetime.strptime(x[0], '%d.%m.%Y')))

    # Save extracted data to MySQL database
    # save_data_to_db(data)

    # Convert the sorted grouped schedules to JSON
    json_data = json.dumps(sorted_grouped_schedules, indent=4, ensure_ascii=False)

    # Save JSON to Cloudinary
    try:
        response = cloudinary.uploader.upload(
            BytesIO(json_data.encode('utf-8')),
            resource_type='raw',
            public_id='sorted_grouped_schedules',  # Customize your public ID
            format='json'  # Specify the format as JSON
        )
        print(f"Uploaded to Cloudinary: {response['url']}")
    except Exception as e:
        print(f"Error uploading to Cloudinary: {e}")

    return sorted_grouped_schedules


@app.post("/upload")
async def upload_file(file: UploadFile = File(...)):
    try:
        # Upload file to Cloudinary
        response = cloudinary.uploader.upload(file.file)
        return JSONResponse(content={"url": response['url'], "public_id": response['public_id']})
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/file/{public_id}")
async def get_file(public_id: str):
    try:
        # Fetch file details from Cloudinary
        response = cloudinary.api.resource(public_id)
        return JSONResponse(
            content={"url": response['url'], "public_id": response['public_id'], "format": response['format']})
    except cloudinary.exceptions.NotFound:
        raise HTTPException(status_code=404, detail="File not found")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


# Function to convert ISO 8601 duration to HH:MM format
def convert_iso_duration_to_hhmm(duration: str) -> str:
    seconds = int(duration[2:-1])  # Extract seconds from the duration string
    hours, remainder = divmod(seconds, 3600)  # Calculate hours and remaining seconds
    minutes, _ = divmod(remainder, 60)  # Calculate minutes
    return f"{hours:02}:{minutes:02}"  # Format to HH:MM


@app.get("/api/schedules/retrieve2", response_model=Dict[str, List[Dict[str, Any]]])
async def retrieve_schedules():
    connection = get_db_connection()
    cursor = connection.cursor(dictionary=True)

    try:
        # Query to retrieve all schedules with formatted times
        query = """
              SELECT 
                  text, 
                  DATE_FORMAT(date, '%d.%m.%Y') AS date, 
                  DATE_FORMAT(start_time, '%H:%i') AS start_time, 
                  DATE_FORMAT(end_time, '%H:%i') AS end_time 
              FROM 
                  schedules
              """
        cursor.execute(query)

        # Fetch all results
        results = cursor.fetchall()

        # Group schedules by date
        grouped_schedules = defaultdict(list)
        for schedule in results:
            grouped_schedules[schedule['date']].append({
                'text': schedule['text'],
                'date': schedule['date'],
                'start_time': schedule['start_time'],
                'end_time': schedule['end_time']
            })

        # Sort the grouped schedules by date
        sorted_grouped_schedules = dict(
            sorted(grouped_schedules.items(), key=lambda x: datetime.strptime(x[0], '%d.%m.%Y'))
        )

        return sorted_grouped_schedules

    except Exception as e:
        raise HTTPException(status_code=500, detail=str(e))
    finally:
        cursor.close()
        connection.close()

@app.get("/api/schedules/retrieve", response_model=Dict[str, Any])
async def retrieve_sorted_grouped_schedules():
    public_id = 'sorted_grouped_schedules.json'  # Hardcoded public_id
    try:
        # Fetch file details from Cloudinary
        response = cloudinary.api.resource(public_id, resource_type='raw')

        # Get the URL to download the JSON file
        file_url = response['secure_url']

        # Download the JSON file from the URL
        json_response = requests.get(file_url)

        if json_response.status_code == 200:
            # Return the JSON data
            return json_response.json()
        else:
            raise HTTPException(status_code=404, detail="File not found or could not be retrieved")

    except cloudinary.exceptions.NotFound:
        raise HTTPException(status_code=404, detail="File not found in Cloudinary")
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))

@app.get("/api/schedules/list_files")
async def list_files():
    try:
        resources = cloudinary.api.resources(resource_type='raw')
        return resources['resources']
    except Exception as e:
        raise HTTPException(status_code=400, detail=str(e))


@app.get("/extract-table/")
async def extract_table():
    url = 'https://rudnik.pl/wp-content/uploads/2023/12/R2.pdf'

    try:
        # Download the PDF file from the specified URL
        response = requests.get(url)
        response.raise_for_status()  # Raise an error for bad responses

        extracted_data = []

        # Use pdfplumber to extract tables from the PDF
        pdf = pdfplumber.open(BytesIO(response.content))  # Open PDF directly from bytes

        # Iterate through each page in the PDF
        for page in pdf.pages:
            tables = page.extract_tables()  # Extract tables from the current page
            for table in tables:
                df = pd.DataFrame(table[1:], columns=table[0])  # Create DataFrame
                extracted_data.append(df.to_dict(orient='records'))  # Convert to dict

        pdf.close()  # Close the PDF file when done

        return JSONResponse(extracted_data[1])

    except requests.HTTPError as http_err:
        raise HTTPException(status_code=http_err.response.status_code, detail=str(http_err))
    except Exception as err:
        raise HTTPException(status_code=500, detail=str(err))

    # # Upload JSON data to Cloudinary
    # try:
    #     response = cloudinary.uploader.upload(
    #         BytesIO(json_data.encode('utf-8')),
    #         resource_type='raw',
    #         public_id='extracted_data',  # Customize your public ID
    #         format='json'  # Specify the format as JSON
    #     )
    #     print(f"Uploaded to Cloudinary: {response['url']}")
    # except Exception as e:
    #     print(f"Error uploading to Cloudinary: {e}")
    #     raise HTTPException(status_code=500, detail="Failed to upload extracted data to Cloudinary")
    #
    # return JSONResponse(content={"url": response['url'], "public_id": response['public_id']})
